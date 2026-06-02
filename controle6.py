import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
import datetime
from datetime import datetime
import sqlite3

def cria_pastas():
    os.makedirs('base_convertida', exist_ok=True)
    os.makedirs('base_validada', exist_ok=True)
    os.makedirs('base_sql', exist_ok=True)
    #os.makedirs(r'base_validada\usuarios', exist_ok=True)

def criar_banco_devolutivas():
    conn = sqlite3.connect(r".\base_sql\devolutivas.db")
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS devolutivas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        data_hora TEXT,
        email TEXT,
        nome TEXT,
        setor TEXT,
        projeto TEXT,
        coluna TEXT,
        Inicio_da_Liberacao_DEVOLUTIVA TEXT,
        Fim_de_Liberacao_DEVOLUTIVA TEXT,
        valor TEXT
    )
    """)

    conn.commit()
    conn.close()

def criar_banco_cadastro():
    conn = sqlite3.connect(r".\base_sql\usuarios.db")
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS cadastro (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    setor TEXT,
    evento TEXT,
    nome TEXT,
    email TEXT,
    Grupo0 TEXT,
    Grupo1 TEXT,
    Grupo2 TEXT,
    Grupo3 TEXT,
    Grupo4 TEXT,
    Grupo5 TEXT,
    Grupo6 TEXT,
    Grupo7 TEXT,
    Grupo8 TEXT,
    Grupo9 TEXT,
    Grupo_UNLOCKED TEXT,
    Grupo_Unlocked_Compras TEXT,
    Grupo_Unlocked_Prog_Disp TEXT
    )
    """)
    
    conn2 = sqlite3.connect(r".\base_sql\usuarios.db")
    cur2 = conn.cursor()

    cur2.execute("""
    CREATE TABLE IF NOT EXISTS sessao_ativa (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    email TEXT, 
    nome TEXT, 
    data_login TEXT
    )
    """)

def criar_banco_demandas():
    conn = sqlite3.connect(r".\base_sql\demandas.db")
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS tarefas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    DATA_PROJ TEXT,
    PROGRAMA TEXT,
    PROJECT TEXT,
    PRODUCT TEXT,
    NOVO_CAT TEXT,
    NUM_CAT TEXT,
    FRONT_PAGE TEXT,
    TEVON TEXT,
    KOM TEXT,
    KOM_DEVOLUTIVA TEXT,
    inicio_da_liberacao TEXT,
    Inicio_da_Liberacao_DEVOLUTIVA TEXT,
    Inicio_da_Liberacao_G0 TEXT,
    Inicio_da_Liberacao_G1 TEXT,
    Inicio_da_Liberacao_G2 TEXT,
    Inicio_da_Liberacao_G3 TEXT,
    Inicio_da_Liberacao_G4 TEXT,
    Inicio_da_Liberacao_G5 TEXT,
    Inicio_da_Liberacao_G6 TEXT,
    Inicio_da_Liberacao_G7 TEXT,
    Inicio_da_Liberacao_G8 TEXT,
    Inicio_da_Liberacao_G9 TEXT,
    Inicio_de_Compras TEXT,
    Inicio_de_Compras_DEVOLUTIVA TEXT,
    Fim_de_Liberacao TEXT,
    Fim_de_Liberacao_DEVOLUTIVA TEXT,
    Fim_de_Liberacao_DEVOLUTIVA_G0 TEXT,
    Fim_de_Liberacao_DEVOLUTIVA_G1 TEXT,     
    Fim_de_Liberacao_DEVOLUTIVA_G2 TEXT,     
    Fim_de_Liberacao_DEVOLUTIVA_G3 TEXT,     
    Fim_de_Liberacao_DEVOLUTIVA_G4 TEXT,     
    Fim_de_Liberacao_DEVOLUTIVA_G5 TEXT,     
    Fim_de_Liberacao_DEVOLUTIVA_G6 TEXT,     
    Fim_de_Liberacao_DEVOLUTIVA_G7 TEXT,     
    Fim_de_Liberacao_DEVOLUTIVA_G8 TEXT,     
    Fim_de_Liberacao_DEVOLUTIVA_G9 TEXT,
    Fim_de_Compras TEXT,
    Fim_de_Compras_DEVOLUTIVA TEXT,
    Início_de_Programacao_Disposicao TEXT,
    Início_de_Programacao_Disposicao__DEVOLUTIVA TEXT,
    Fim_da_Programacao_Disposicao TEXT,
    Fim_da_Programacao_Disposicao_DEVOLUTIVA TEXT,
    SOP TEXT,
    ME TEXT,
    COMMENTS TEXT
    )
    """)

def paleta():
    leitor = pd.read_excel(r'paleta_cores.xlsx', sheet_name='paleta de cores', usecols=range(3))
    leitor.to_excel(r'.\base_convertida\paleta_cores.xlsx', index=False)

def controle():
    leitor = pd.read_excel('Project Summary - Volkswagen.xlsx', sheet_name='PROGRAMAS')
    leitor = leitor.drop(columns=leitor.columns[0])
    leitor.to_excel(r'.\base_convertida\controle.xlsx', index=False)

def flag():
    leitor = pd.read_excel('Project Summary - Volkswagen.xlsx', sheet_name='Flag')
    #leitor = leitor.drop(columns=leitor.columns[0])
    leitor.to_excel(r'.\base_convertida\flag.xlsx', index=False)

def ws_pep():
    leitor = pd.read_excel('Project Summary - Volkswagen.xlsx', sheet_name='PEP')
    #leitor = leitor.drop(columns=leitor.columns[0])
    leitor['Unidade de medida'] = 'Dias'
    leitor.to_excel(r'.\base_convertida\pep.xlsx', index=False)

def validacao():
    leitor = pd.read_excel(r'.\base_convertida\controle.xlsx',skiprows=3)
    leitor.to_excel(r'.\base_validada\validacao_controle_TABELA.xlsx', index=False, header=True)

def tabela_padrao():
    leitor = pd.read_excel(r'.\base_validada\validacao_controle_TABELA.xlsx')
    with pd.ExcelWriter(r'.\base_validada\validacao_controle_TABELA.xlsx', engine="openpyxl", mode='w') as writer:
        book = writer.book
        leitor.to_excel(writer, index=False, sheet_name='Controle do Projeto')
        ws = book['Controle do Projeto']
        ws['A1'] = 'DATA'
        ws['B1'] = 'PROGRAMA'
        ws['C1'] = 'PROJECT'
        ws['D1'] = 'PRODUCT'
        ws['E1'] = 'NOVO CATÁLOGO'
        ws['F1'] = 'NÚMERO CATÁLOGO'
        ws['G1'] = 'FRONT PAGE'
        ws['H1'] = 'TEVON'
        ws['I1'] = 'KOM'
        ws['J1'] = 'inicio_da_liberacao'
        ws['K1'] = 'Início de Compras'
        ws['L1'] = 'Fim de Liberação'
        ws['M1'] = 'Fim de Compras'
        ws['N1'] = 'Início de Programação/Disposição'
        ws['O1'] = 'Fim da Programação/Disposição'
        ws['P1'] = 'SOP'
        ws['Q1'] = 'ME'
        ws['R1'] = 'COMMENTS:'

def validacao_6m():
    leitor = pd.read_excel(r'.\base_validada\validacao_controle_TABELA.xlsx')
    leitor2 = leitor[leitor['DATA'] == '6 MESES']
    leitor2.to_excel(r'.\base_validada\validacao_controle_TABELA_6MESES.xlsx', index=False, header=True)
        
def validacao_posterior():
    leitor = pd.read_excel(r'.\base_validada\validacao_controle_TABELA.xlsx')
    leitor2 = leitor[leitor['DATA'] == 'POSTERIOR']
    leitor2.to_excel(r'.\base_validada\validacao_controle_TABELA_POSTERIOR.xlsx', index=False, header=True)

def validacao_vazio():
    leitor = pd.read_excel(r'.\base_validada\validacao_controle_TABELA.xlsx')
    leitor2 = leitor[~leitor['DATA'].isin(['6 MESES', 'POSTERIOR'])]
    leitor2.to_excel(r'.\base_validada\validacao_controle_TABELA_VAZIO.xlsx', index=False, header=True)

def compilacao():
    leitor = pd.read_excel(r'.\base_validada\validacao_controle_TABELA.xlsx')
    leitor2 = pd.read_excel(r'.\base_validada\validacao_controle_TABELA_6MESES.xlsx')
    leitor3 = pd.read_excel(r'.\base_validada\validacao_controle_TABELA_POSTERIOR.xlsx')
    leitor4 = pd.read_excel(r'.\base_validada\validacao_controle_TABELA_VAZIO.xlsx')
    leitor5 = pd.read_excel(r'.\base_convertida\flag.xlsx')
    leitor6 = pd.read_excel(r'.\base_convertida\pep.xlsx')
    leitor7 = pd.read_excel(r'.\base_convertida\paleta_cores.xlsx')
    with pd.ExcelWriter(r'.\base_validada\base_compilada.xlsx', engine='xlsxwriter', date_format='dd/mm/yyyy') as writer:
        leitor.to_excel(writer, index=False, sheet_name='Controle do Projeto')
        leitor2.to_excel(writer, index=False, sheet_name='6 MESES')
        leitor3.to_excel(writer, index=False, sheet_name='POSTERIOR')
        leitor4.to_excel(writer, index=False, sheet_name='SEM PERÍODO')
        leitor5.to_excel(writer, index=False, sheet_name='FLAG')
        leitor6.to_excel(writer, index=False, sheet_name='PEP')
        leitor7.to_excel(writer, index=False, sheet_name='PALETA_DE_CORES')
        pd.DataFrame().to_excel(writer, index=False, sheet_name='LOG')

def semaforo():
    wb = load_workbook(r'.\base_validada\base_compilada.xlsx')
    ws = wb['Controle do Projeto']
    ws.insert_cols(10); ws['J1'] = 'KOM - DEVOLUTIVA'
    ws.insert_cols(12); ws['L1'] = 'Inicio_da_Liberacao_DEVOLUTIVA'
    ws.insert_cols(13); ws['M1'] = 'Início da Liberação - DEVOLUTIVA - G0'
    ws.insert_cols(14); ws['N1'] = 'Início da Liberação - DEVOLUTIVA - G1'
    ws.insert_cols(15); ws['O1'] = 'Início da Liberação - DEVOLUTIVA - G2'
    ws.insert_cols(16); ws['P1'] = 'Início da Liberação - DEVOLUTIVA - G3'
    ws.insert_cols(17); ws['Q1'] = 'Início da Liberação - DEVOLUTIVA - G4'
    ws.insert_cols(18); ws['R1'] = 'Início da Liberação - DEVOLUTIVA - G5'
    ws.insert_cols(19); ws['S1'] = 'Início da Liberação - DEVOLUTIVA - G6'
    ws.insert_cols(20); ws['T1'] = 'Início da Liberação - DEVOLUTIVA - G7'
    ws.insert_cols(21); ws['U1'] = 'Início da Liberação - DEVOLUTIVA - G8'
    ws.insert_cols(22); ws['V1'] = 'Início da Liberação - DEVOLUTIVA - G9'
    ws.insert_cols(24); ws['X1'] = 'Início de Compras - DEVOLUTIVA'
    ws.insert_cols(26); ws['Z1'] = 'Fim de Liberação - DEVOLUTIVA'
    ws.insert_cols(27); ws['AA1'] = 'Fim da Liberação - DEVOLUTIVA - G0'
    ws.insert_cols(28); ws['AB1'] = 'Fim da Liberação - DEVOLUTIVA - G1'
    ws.insert_cols(29); ws['AC1'] = 'Fim da Liberação - DEVOLUTIVA - G2'
    ws.insert_cols(30); ws['AD1'] = 'Fim da Liberação - DEVOLUTIVA - G3'
    ws.insert_cols(31); ws['AE1'] = 'Fim da Liberação - DEVOLUTIVA - G4'
    ws.insert_cols(32); ws['AF1'] = 'Fim da Liberação - DEVOLUTIVA - G5'
    ws.insert_cols(33); ws['AG1'] = 'Fim da Liberação - DEVOLUTIVA - G6'
    ws.insert_cols(34); ws['AH1'] = 'Fim da Liberação - DEVOLUTIVA - G7'
    ws.insert_cols(35); ws['AI1'] = 'Fim da Liberação - DEVOLUTIVA - G8'
    ws.insert_cols(36); ws['AJ1'] = 'Fim da Liberação - DEVOLUTIVA - G9'
    ws.insert_cols(38); ws['AL1'] = 'Fim de Compras - DEVOLUTIVA'
    ws.insert_cols(40); ws['AN1'] = 'Início de Programação/Disposição - DEVOLUTIVA'
    ws.insert_cols(42); ws['AP1'] = 'Fim da Programação/Disposição - DEVOLUTIVA'
    wb.save(r'.\base_validada\base_compilada_padrao.xlsx')

def importador_flag():
    lista_setor = []
    lista_evento = []
    lista_nome = []
    lista_email = []
    lista_grupo1 = []
    lista_grupo2 = []
    lista_grupo3 = []
    lista_grupo4 = []
    lista_grupo5 = []
    lista_grupo6 = []
    lista_grupo7 = []
    lista_grupo8 = []
    lista_grupo9 = []
    lista_grupo10 = []
    lista_grupo11 = []
    lista_grupo12 = []
    lista_grupo13 = []

    leitor = pd.read_excel(r'.\base_validada\base_compilada_padrao.xlsx', sheet_name='FLAG')
    conta_linha = len(leitor)
    conta_colunas = len(leitor.columns)
    
    for a in range(conta_linha):
        lista_evento.append(str(leitor.iloc[a,1]))
        lista_nome.append(str(leitor.iloc[a,2]))
        lista_email.append(str(leitor.iloc[a,3]))
        lista_setor.append(str(leitor.iloc[a,0]))
        lista_grupo1.append(str(leitor.iloc[a,4]))
        lista_grupo2.append(str(leitor.iloc[a,5]))
        lista_grupo3.append(str(leitor.iloc[a,6]))
        lista_grupo4.append(str(leitor.iloc[a,7]))
        lista_grupo5.append(str(leitor.iloc[a,8]))
        lista_grupo6.append(str(leitor.iloc[a,9]))
        lista_grupo7.append(str(leitor.iloc[a,10]))
        lista_grupo8.append(str(leitor.iloc[a,11]))
        lista_grupo9.append(str(leitor.iloc[a,12]))
        lista_grupo10.append(str(leitor.iloc[a,13]))
        lista_grupo11.append(str(leitor.iloc[a,14]))
        lista_grupo12.append(str(leitor.iloc[a,15]))
        lista_grupo13.append(str(leitor.iloc[a,16]))
        
# ---------- CONEXÃO COM SQLITE ----------
    conexao = sqlite3.connect(r".\base_sql\usuarios.db")
    cursor = conexao.cursor()


# ---------- QUERY ----------
    sql = """
    INSERT OR IGNORE INTO cadastro (
    setor,
    evento,
    nome,
    email,
    Grupo0,
    Grupo1,
    Grupo2,
    Grupo3,
    Grupo4,
    Grupo5,
    Grupo6,
    Grupo7,
    Grupo8,
    Grupo9,
    Grupo_UNLOCKED,
    Grupo_Unlocked_Compras,
    Grupo_Unlocked_Prog_Disp
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """

    # ---------- LOOP DE INSERT ----------
    for b in range(conta_linha):
        valores = (
            lista_setor[b],
            lista_evento[b],
            lista_nome[b],
            lista_email[b],          
            lista_grupo1[b],
            lista_grupo2[b],
            lista_grupo3[b],
            lista_grupo4[b],
            lista_grupo5[b],
            lista_grupo6[b],
            lista_grupo7[b],
            lista_grupo8[b],
            lista_grupo9[b],
            lista_grupo10[b],
            lista_grupo11[b],
            lista_grupo12[b],
            lista_grupo13[b]

         )

        cursor.execute(sql, valores)

    # ---------- FINALIZA ----------
    conexao.commit()
    conexao.close()

def formatar_data(valor):
    if pd.isna(valor):
        return None
    try:
        return valor.strftime('%d/%m/%Y')
    except:
        return None

def importador_demandas():
    leitor = pd.read_excel(
        r'.\base_validada\base_compilada_padrao.xlsx',
        sheet_name='Controle do Projeto'
    )

    conta_linhas = len(leitor)
    #print(f'Nº de linhas {conta_linhas}')
    #print(f'Nº de colunas {len(leitor.columns)}')
    #print('-' * 100)

    # ---------- CONEXÃO ----------
    conexao = sqlite3.connect(r".\base_sql\demandas.db")
    cursor = conexao.cursor()

    # ---------- QUERY CORRIGIDA ----------
    query = """
    INSERT OR IGNORE INTO tarefas (
        DATA_PROJ, 
        PROGRAMA, 
        PROJECT, 
        PRODUCT, 
        NOVO_CAT, 
        NUM_CAT,
        FRONT_PAGE, 
        TEVON, 
        KOM, 
        KOM_DEVOLUTIVA,
        inicio_da_liberacao,
        Inicio_da_Liberacao_DEVOLUTIVA,
        Inicio_da_Liberacao_G0,
        Inicio_da_Liberacao_G1,
        Inicio_da_Liberacao_G2,
        Inicio_da_Liberacao_G3,
        Inicio_da_Liberacao_G4,
        Inicio_da_Liberacao_G5,
        Inicio_da_Liberacao_G6,
        Inicio_da_Liberacao_G7,
        Inicio_da_Liberacao_G8,
        Inicio_da_Liberacao_G9,
        InIcio_de_Compras,
        Inicio_de_Compras_DEVOLUTIVA,
        Fim_de_Liberacao,
        Fim_de_Liberacao_DEVOLUTIVA,
        Fim_de_Liberacao_DEVOLUTIVA_G0,
        Fim_de_Liberacao_DEVOLUTIVA_G1,     
        Fim_de_Liberacao_DEVOLUTIVA_G2,     
        Fim_de_Liberacao_DEVOLUTIVA_G3,     
        Fim_de_Liberacao_DEVOLUTIVA_G4,     
        Fim_de_Liberacao_DEVOLUTIVA_G5,     
        Fim_de_Liberacao_DEVOLUTIVA_G6,     
        Fim_de_Liberacao_DEVOLUTIVA_G7,     
        Fim_de_Liberacao_DEVOLUTIVA_G8,     
        Fim_de_Liberacao_DEVOLUTIVA_G9,
        Fim_de_Compras,
        Fim_de_Compras_DEVOLUTIVA,
        Início_de_Programacao_Disposicao,
        Início_de_Programacao_Disposicao__DEVOLUTIVA,
        Fim_da_Programacao_Disposicao,
        Fim_da_Programacao_Disposicao_DEVOLUTIVA,
        SOP,
        ME,
        COMMENTS
        )
    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """

    # ---------- LOOP ----------
    for i in range(conta_linhas):
        valores = (
            leitor.iloc[i, 0],
            leitor.iloc[i, 1],
            leitor.iloc[i, 2],
            leitor.iloc[i, 3],
            leitor.iloc[i, 4],
            leitor.iloc[i, 5],
            leitor.iloc[i, 6],
            leitor.iloc[i, 7],
            formatar_data(leitor.iloc[i, 8]),
            leitor.iloc[i, 9],
            formatar_data(leitor.iloc[i, 10]),
            leitor.iloc[i, 11],
            leitor.iloc[i, 12],
            leitor.iloc[i, 13],
            leitor.iloc[i, 14],
            leitor.iloc[i, 15],
            leitor.iloc[i, 16],
            leitor.iloc[i, 17],
            leitor.iloc[i, 18],
            leitor.iloc[i, 19],
            leitor.iloc[i, 20],
            leitor.iloc[i, 21],
            formatar_data(leitor.iloc[i, 22]),
            leitor.iloc[i, 23],
            formatar_data(leitor.iloc[i, 24]),
            leitor.iloc[i,25],
            leitor.iloc[i,26],
            leitor.iloc[i,27],
            leitor.iloc[i,28],
            leitor.iloc[i,29],
            leitor.iloc[i,30],
            leitor.iloc[i,31],
            leitor.iloc[i,32],
            leitor.iloc[i,33],
            leitor.iloc[i,34],
            leitor.iloc[i,35],
            formatar_data(leitor.iloc[i,36]),
            leitor.iloc[i,37],
            formatar_data(leitor.iloc[i,38]),
            leitor.iloc[i,39],
            formatar_data(leitor.iloc[i,40]),
            leitor.iloc[i,41],
            formatar_data(leitor.iloc[i,42]),
            formatar_data(leitor.iloc[i,43]),
            leitor.iloc[i,44]
            #'Inicio_da_Liberacao_G0',
            #leitor.iloc[i, 13],
            #leitor.iloc[i, 14],
            #formatar_data(leitor.iloc[i, 15]),
            #leitor.iloc[i, 16],
            #formatar_data(leitor.iloc[i, 17]),
            #leitor.iloc[i, 18],
            #formatar_data(leitor.iloc[i, 19]),
            #leitor.iloc[i, 20],
            #formatar_data(leitor.iloc[i, 21]),
            #leitor.iloc[i, 22],
            #formatar_data(leitor.iloc[i, 23]),
            #formatar_data(leitor.iloc[i, 24]),
            #leitor.iloc[i, 25],
        )

        cursor.execute(query, valores)

    conexao.commit()
    conexao.close()
    

os.system('cls')
cria_pastas()
print('Criação de pastas - OK!')
criar_banco_devolutivas()
print('Criação de banco de dados devolutivas - OK!')
criar_banco_cadastro()
print('Criação de banco de dados cadastro - OK!')
criar_banco_demandas()
print('Criação de banco de dados demandas - OK!')
paleta()
controle()
flag()
ws_pep()
validacao()
tabela_padrao()
validacao_6m()
validacao_posterior()
validacao_vazio()
compilacao()
semaforo()
print('Criação de base compilada - OK!')
importador_flag()
print('Importação de base Flag - OK!')
importador_demandas()
print('Importação de base Demandas - OK!')
print('{:.<100}'.format(''))